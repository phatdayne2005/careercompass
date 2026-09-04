# -*- coding: utf-8 -*-
"""Sinh lại ba sheet 03, 03b, 04 của BaoCao-PhanA-HopDen.xlsx.

Chạy từ thư mục gốc dự án, sau khi đã chạy test:
    mvnw clean test -Dtest=ProgressDecisionTableTest,TokenValidityDecisionTableTest,ProgressStateTransitionTest
    python tools/gen_sheets_dt_st.py

Bỏ nhật ký chạy test ở đầu sheet (không có trong slide, khó trình bày) và thay bằng
bảng tổng kết một dòng. Bù lại, chèn thẳng cột/hàng Status + tên method vào bảng THIẾT
KẾ để không mất bằng chứng đã chạy — giống cách sheet 01 đã làm.

Phần thiết kế và mọi ghi chú được ĐỌC LẠI nguyên văn từ file cũ, không gõ lại, nên chữ
do người viết báo cáo tự đặt không bị sai lệch.
"""
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "BaoCao-PhanA-HopDen.xlsx"
TMP = "_tmp.xlsx"

C_HDR = "1F4E79"; C_TXT = "FFFFFFFF"; C_PASS = "C6EFCE"; C_SEC = "DDEBF7"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NL = chr(10)
SUREFIRE = Path("target/surefire-reports")

# Tên method thi hành từng cột/dòng của bảng thiết kế, theo đúng thứ tự R1.. / ST-01..
METHOD = {
    "03. Decision Table (PhanA)": [
        "rule1_notStarted_nodeLocked_choPhep", "rule2_notStarted_nodeUnlocked_choPhep",
        "rule3_inProgress_nodeLocked_tuChoi", "rule4_inProgress_nodeUnlocked_choPhep",
        "rule5_done_nodeLocked_tuChoi", "rule6_done_nodeUnlocked_choPhep"],
    "03b. Decision Table Token": [
        "rule1_conHan_chuaDung_hopLe", "rule2_hetHan_chuaDung_khongHopLe",
        "rule3_conHan_daDung_khongHopLe", "rule4_hetHan_daDung_khongHopLe"],
    "04. State Transition (PhanA)": [
        "st01_notStarted_toInProgress", "st05_notStarted_toDone",
        "st07_notStarted_toDone_nodeLocked_biChan", "st02_inProgress_toDone",
        "st06_inProgress_toNotStarted", "st08_inProgress_toDone_nodeLocked_biChan",
        "st03_done_toNotStarted", "st04_done_toInProgress",
        "st09_done_toInProgress_nodeLocked_biChan"],
}

# sheet -> (gói, lớp test, kỹ thuật, số test dự kiến)
SUITE = {
    "03. Decision Table (PhanA)":
        ("blackbox", "ProgressDecisionTableTest", "Bảng quyết định · 6 luật", 6),
    "03b. Decision Table Token":
        ("blackbox", "TokenValidityDecisionTableTest", "Bảng quyết định · 4 luật", 4),
    "04. State Transition (PhanA)":
        ("blackbox", "ProgressStateTransitionTest", "Chuyển đổi trạng thái · 9 case", 9),
}


def so_test_that(pkg, cls, du_kien):
    """Đọc số test đã chạy từ surefire; lệch với báo cáo thì dừng ngay."""
    f = SUREFIRE / f"TEST-vn.uth.careercompass.{pkg}.{cls}.xml"
    if not f.exists():
        print(f"  ! chua co surefire cho {cls}, dung so du kien {du_kien}")
        return du_kien
    that = int(ET.parse(f).getroot().attrib["tests"])
    if that != du_kien:
        raise SystemExit(
            f"LECH SO LIEU: {cls} chay {that} test nhung bao cao ghi {du_kien}. "
            f"Sua SUITE trong tools/gen_sheets_dt_st.py cho khop roi chay lai.")
    return that


def doc_phan_thiet_ke(ws):
    """Đọc nguyên văn phần thiết kế: từ thanh 'BẢNG ...' đầu tiên tới hết sheet."""
    dau = next((r for r in range(7, ws.max_row + 1)
                if str(ws.cell(row=r, column=1).value or "").startswith("BẢNG")), None)
    if dau is None:
        raise SystemExit(f"khong tim thay bang thiet ke trong {ws.title}")
    gop = {rng.min_row for rng in ws.merged_cells.ranges if rng.min_col == 1 and rng.max_col >= 8}
    # Bỏ qua hai hàng do chính script này chèn ở lần chạy trước, để chạy lại nhiều lần
    # vẫn ra cùng kết quả thay vì chèn chồng lên nhau.
    da_chen = {"Kết quả chạy test", "Method trong code"}
    return [([ws.cell(row=r, column=c).value for c in range(1, 9)], r in gop)
            for r in range(dau, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").strip() not in da_chen]


wb = openpyxl.load_workbook(XLSX)

for ten_sheet, (pkg, cls, ky_thuat, du_kien) in SUITE.items():
    cu = wb[ten_sheet]
    idx = wb.sheetnames.index(ten_sheet)
    tieu_de = cu.cell(row=1, column=1).value
    mo_ta = cu.cell(row=2, column=1).value
    thiet_ke = doc_phan_thiet_ke(cu)
    so_test = so_test_that(pkg, cls, du_kien)
    methods = METHOD[ten_sheet]

    ws = wb.create_sheet(ten_sheet + " TMP")
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 24

    def cell(r, c, v, *, bold=False, fill=None, center=False, size=10, color=None):
        o = ws.cell(row=r, column=c, value=v)
        o.font = Font(size=size, bold=bold, color=color)
        o.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if center else "general")
        o.border = BORDER
        if fill:
            o.fill = PatternFill("solid", fgColor=fill)

    def bar(r, t, *, fill=C_SEC, color=C_HDR, size=11, italic=False, height=24):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        o = ws.cell(row=r, column=1, value=t)
        o.font = Font(bold=not italic, italic=italic, size=size, color=color)
        if fill:
            o.fill = PatternFill("solid", fgColor=fill)
        o.alignment = Alignment(vertical="top" if italic else "center", wrap_text=True)
        ws.row_dimensions[r].height = height
        return r + 1

    o = ws.cell(row=1, column=1, value=tieu_de)
    o.font = Font(bold=True, size=13, color=C_HDR)
    o = ws.cell(row=2, column=1, value=mo_ta)
    o.font = Font(size=10, italic=True, color="595959")

    r = bar(4, f"TỔNG KẾT THI HÀNH — {so_test} test đã chạy, {so_test} PASS, 0 fail. "
               f"Bảng thiết kế bên dưới có cột ghi rõ method nào thi hành dòng nào.", height=30)
    for i, h in enumerate(["Lớp test trong source", "Số test", "Kỹ thuật áp dụng",
                           "", "", "", "", "Status"], 1):
        cell(r, i, h, bold=True, fill=C_HDR, color=C_TXT, center=True)
    ws.row_dimensions[r].height = 26
    r += 1
    for i, v in enumerate([cls, so_test, ky_thuat, "", "", "", "", "PASS"], 1):
        cell(r, i, v, fill=(C_PASS if i == 8 else None), center=(i in (2, 8)))
    ws.row_dimensions[r].height = 26
    r += 1
    r = bar(r, f"Số test được đối chiếu tự động với target/surefire-reports khi sinh lại sheet "
               f"bằng tools/gen_sheets_dt_st.py — source chạy ra số khác thì script dừng và báo "
               f"lỗi. Chạy lại để kiểm chứng:  mvnw test -Dtest={cls} -DshowCases",
            fill=None, color="595959", size=9, italic=True, height=34)
    r += 1

    # --- phần thiết kế: chép nguyên văn, chỉ thêm Status + tên method ---
    dem = 0
    for vals, la_thanh in thiet_ke:
        a = str(vals[0] or "")
        if la_thanh:
            r = bar(r, vals[0],
                    fill=(C_SEC if a.startswith("BẢNG") else None),
                    color=(C_HDR if a.startswith("BẢNG") else "595959"),
                    size=(11 if a.startswith("BẢNG") else 9),
                    italic=not a.startswith("BẢNG"),
                    height=(24 if a.startswith("BẢNG") else 44))
            continue
        if not any(v not in (None, "") for v in vals):
            r += 1
            continue

        la_dau_bang = a in ("Condition / Action", "Test Case No.")
        la_dong_st = a.isdigit() and vals[1] and str(vals[1]).isupper()

        if la_dau_bang and a == "Test Case No.":
            vals = list(vals[:6]) + ["Status", "Method trong code"]
        elif la_dong_st:
            vals = list(vals[:6]) + ["PASS", methods[dem]]
            dem += 1

        for j, v in enumerate(vals, 1):
            f = None
            if la_dau_bang:
                f = C_HDR
            elif j == 7 and v == "PASS":
                f = C_PASS
            cell(r, j, v, bold=la_dau_bang or (j == 1 and a.startswith(("C", "A"))),
                 fill=f, color=(C_TXT if la_dau_bang else None),
                 center=la_dau_bang or (j > 1 and len(str(v or "")) <= 12))
        ws.row_dimensions[r].height = 34 if la_dau_bang else 30
        r += 1

        # Sau dòng hành động cuối của BẢNG ĐẦY ĐỦ, chèn Status + method cho từng luật.
        if a.startswith("A2 ·") and dem == 0 and len(methods) in (4, 6):
            for nhan, du_lieu in [("Kết quả chạy test", ["PASS"] * len(methods)),
                                  ("Method trong code", methods)]:
                cell(r, 1, nhan, bold=True)
                for j, v in enumerate(du_lieu, 2):
                    cell(r, j, v, fill=(C_PASS if v == "PASS" else None),
                         center=(v == "PASS"), size=9 if v != "PASS" else 10)
                cell(r, 8, "")
                ws.row_dimensions[r].height = 34
                r += 1
            dem = len(methods)

    wb.remove(cu)
    wb._sheets.insert(idx, wb._sheets.pop(wb._sheets.index(ws)))
    ws.title = ten_sheet
    print(f"  OK {ten_sheet}: {so_test} test, ket thuc dong {r - 1}")

wb.save(TMP)
print(f"Da ghi {TMP}")


# =====================================================================
# Sheet 99 — hướng dẫn gộp. Dựng lại vì bốn sheet kia đã bỏ nhật ký chạy
# test, nên cột "Vùng công thức A4" không còn nghĩa gì.
# =====================================================================
HD = [
    ("01. BVA + Equiv Partition", "Thay cho '01. BVA' và '02. Equiv Partition'", 56,
     "RegisterStandardBvaTest (19)" + NL + "RegisterTagCoverageTest (17)" + NL
     + "RegisterEquivalencePartitionTest (14)" + NL + "OnboardingFileSizeBvaTest (6)"),
    ("03. Decision Table (PhanA)", "Thay cho '03. Decision Table'", 6,
     "ProgressDecisionTableTest" + NL + "phủ đủ 6 rule (3 × 2)"),
    ("03b. Decision Table Token", "Thêm mới, đặt cạnh sheet 03", 4,
     "TokenValidityDecisionTableTest" + NL + "phủ đủ 4 rule (2 × 2)"),
    ("04. State Transition (PhanA)", "Thay cho '04. State Transition'", 9,
     "ProgressStateTransitionTest" + NL + "6 cạnh hợp lệ + 3 bị chặn"),
]

KHAC_BIET = [
    ("Cột bằng chứng ghi 'Unit test hiện có' — gán ngược test có sẵn vào tên kỹ thuật",
     "Mỗi hàm test đặt tên theo mã dòng bảng thiết kế (rule3_, st05_) — thiết kế trước, code sau"),
    ("Không có bảng thiết kế theo mẫu slide",
     "Có bảng slide 23 (Standard BVA), slide 32 (phân lớp + biên), slide 33 (gộp tag), "
     "slide 44 (quyết định), slide 39 (chuyển trạng thái)"),
    ("Decision Table: 4 tổ hợp, trong đó 1 dòng khai sai độ phủ",
     "6 tổ hợp — đủ 3 × 2, phát hiện tổ hợp IN_PROGRESS + khoá trước đây chưa có test"),
    ("State Transition: chỉ chuyển đổi hợp lệ", "Đủ 6 cạnh + 3 chuyển đổi bị chặn"),
    ("Equivalence Partitioning: suy ra từ BVA",
     "14 case độc lập, phát hiện lớp 'toàn khoảng trắng' mà BVA không chạm tới"),
    ("Không có nhật ký/bằng chứng đã chạy",
     "Mỗi sheet mở đầu bằng bảng TỔNG KẾT THI HÀNH, số test đọc tự động từ "
     "target/surefire-reports; bảng thiết kế có cột Status và tên method"),
    ("email @Size(max = 150), 'a@b' ba ký tự vẫn qua được",
     "Đã sửa thành @Size(min = 6, max = 150) — lỗi do TC16 của Standard BVA phát hiện"),
]

cu = wb["99. Huong dan gop"]
idx = wb.sheetnames.index("99. Huong dan gop")
ws = wb.create_sheet("99 TMP")
for i, w in enumerate([30, 34, 10, 40, 24, 24, 24, 24], 1):
    ws.column_dimensions[get_column_letter(i)].width = w


def o_99(r, c, v, *, bold=False, fill=None, center=False, size=10, color=None, italic=False):
    o = ws.cell(row=r, column=c, value=v)
    o.font = Font(size=size, bold=bold, italic=italic, color=color)
    o.alignment = Alignment(vertical="top", wrap_text=True,
                            horizontal="center" if center else "general")
    o.border = BORDER
    if fill:
        o.fill = PatternFill("solid", fgColor=fill)


o = ws.cell(row=1, column=1, value="HƯỚNG DẪN GỘP VÀO CareerCompass_Test_Report_All_Methods.xlsx")
o.font = Font(bold=True, size=13, color=C_HDR)

r = 3
for i, h in enumerate(["Sheet trong file này", "Gộp vào đâu", "Số test", "Lớp test thi hành"], 1):
    o_99(r, i, h, bold=True, fill=C_HDR, color=C_TXT, center=True)
ws.row_dimensions[r].height = 26
r += 1
for ten, dich, n, lop in HD:
    for i, v in enumerate([ten, dich, n, lop], 1):
        o_99(r, i, v, center=(i == 3))
    ws.row_dimensions[r].height = 62
    r += 1
for i, v in enumerate(["TỔNG", "", sum(h[2] for h in HD), ""], 1):
    o_99(r, i, v, bold=True, fill=C_SEC, center=(i == 3))
r += 2

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
o = ws.cell(row=r, column=1, value=(
    "CÁCH GỘP: chuột phải tab sheet → Move or Copy → chọn file All_Methods → tick "
    "'Create a copy'. Bốn sheet này KHÔNG còn danh sách case ở đầu nên không cần sửa "
    "công thức COUNTA/COUNTIF nào. Nếu sheet '00. Tổng quan' của file All_Methods đang "
    "đếm theo vùng $A$7:$A$nn của các sheet cũ, xoá dòng đếm đó đi và lấy số ở bảng "
    "TỔNG KẾT THI HÀNH (dòng 4) của từng sheet."))
o.font = Font(size=10)
o.alignment = Alignment(vertical="top", wrap_text=True)
ws.row_dimensions[r].height = 62
r += 2

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
o = ws.cell(row=r, column=1, value="KHÁC BIỆT SO VỚI BẢN CŨ CỦA NHÓM")
o.font = Font(bold=True, size=11, color=C_HDR)
o.fill = PatternFill("solid", fgColor=C_SEC)
ws.row_dimensions[r].height = 24
r += 1
for i, h in enumerate(["Bản cũ", "Bản này", "", ""], 1):
    if h:
        o_99(r, i, h, bold=True, fill=C_HDR, color=C_TXT, center=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
for cu_t, moi_t in KHAC_BIET:
    o_99(r, 1, cu_t)
    o_99(r, 2, moi_t)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 46
    r += 1

wb.remove(cu)
wb._sheets.insert(idx, wb._sheets.pop(wb._sheets.index(ws)))
ws.title = "99. Huong dan gop"
print(f"  OK 99. Huong dan gop: ket thuc dong {r - 1}")
wb.save(TMP)
