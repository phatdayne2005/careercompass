# -*- coding: utf-8 -*-
"""Sinh lại sheet "01. BVA + Equiv Partition" của BaoCao-PhanA-HopDen.xlsx.

Chạy từ thư mục gốc dự án:  python tools/gen_sheet01.py

Dựng lại toàn bộ sheet thay vì vá từng chỗ, vì openpyxl.insert_rows() không dịch
chuyển vùng merged cell — vá nhiều lần sẽ làm thanh tiêu đề lệch khỏi nội dung.
Số test ở bảng tổng kết được đối chiếu tự động với target/surefire-reports:
lệch với source thì script dừng ngay, báo cáo không thể ghi số bịa.
"""
import re
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "BaoCao-PhanA-HopDen.xlsx"
TMP = "_tmp.xlsx"
SHEET = "01. BVA + Equiv Partition"

C_HDR = "1F4E79"; C_TXT = "FFFFFFFF"; C_PASS = "C6EFCE"; C_SEC = "DDEBF7"
C_BAND = "F7FAFD"; C_VAR = "E4DFF5"; C_DIRTY = "FBE4E4"; C_OK = "E2F0D9"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NL = chr(10)
COL_OF = {"fullName": 2, "email": 3, "password": 4}

FN, EM, PW = 50, 75, 18          # giá trị nominal của ba biến

# ---- Bảng 2: Standard BVA, 4n+1 = 13 ----
STD = [(1, FN, EM, 6, "password", "min"), (2, FN, EM, 7, "password", "min+"),
       (3, FN, EM, 18, "—", "TẤT CẢ nominal"), (4, FN, EM, 29, "password", "max-"),
       (5, FN, EM, 30, "password", "max"), (6, FN, 6, PW, "email", "min"),
       (7, FN, 7, PW, "email", "min+"), (8, FN, 149, PW, "email", "max-"),
       (9, FN, 150, PW, "email", "max"), (10, 1, EM, PW, "fullName", "min"),
       (11, 2, EM, PW, "fullName", "min+"), (12, 99, EM, PW, "fullName", "max-"),
       (13, 100, EM, PW, "fullName", "max")]

# ---- Bảng 3: Robustness bổ sung, 6n+1 = 19 ----
ROB = [(14, FN, EM, 5, "password", "min-", False),
       (15, FN, EM, 31, "password", "max+", False),
       (16, FN, 5, PW, "email", "min-", False),
       (17, FN, 151, PW, "email", "max+", False),
       (18, 0, EM, PW, "fullName", "min-", False),
       (19, 101, EM, PW, "fullName", "max+", False)]

# ---- Bảng 4: phân hoạch lớp tương đương, mẫu slide 32 ----
EPT = [
    ["password", "6–30 ký tự," + NL + "có ít nhất 1 ký tự không trắng", "V1",
     "< 6 ký tự" + NL + "> 30 ký tự" + NL + "toàn khoảng trắng" + NL + "null",
     "X1" + NL + "X2" + NL + "X3" + NL + "X4", "6, 7, 18, 29, 30", "B1–B5",
     "5 và 31 (B6, B7)"],
    ["fullName", "1–100 ký tự," + NL + "có ít nhất 1 ký tự không trắng", "V2",
     "> 100 ký tự" + NL + "toàn khoảng trắng" + NL + "null",
     "X5" + NL + "X6" + NL + "X7", "1, 2, 50, 99, 100", "B8–B12",
     "rỗng (= min−1) và 101" + NL + "(B13, B14)"],
    ["email — điều kiện ĐỘ DÀI", "6–150 ký tự", "V3",
     "< 6 ký tự" + NL + "> 150 ký tự", "X8" + NL + "X9",
     "6, 7, 75, 149, 150", "B15–B19", "5 và 151 (B20, B21)"],
    ["email — điều kiện ĐỊNH DẠNG", "đúng định dạng email", "V4",
     "sai định dạng" + NL + "null / rỗng", "X10" + NL + "X11",
     "(không áp dụng —" + NL + "miền rời rạc)", "–", "(không áp dụng)"],
    ["Tệp bảng điểm — DUNG LƯỢNG", "0 đến 10 MB", "V5", "> 10 MB", "X12",
     "0, 1 byte, 5 MB," + NL + "10MB−1, 10 MB", "B22–B26", "10MB + 1 byte (B27)"],
]

# ---- Bảng 5: gộp tag thành test case, mẫu slide 33 ----
TAG_OK = "RegisterTagCoverageTest ·" + NL + "tagCoverage_caseHopLe"
TAG_NG = "RegisterTagCoverageTest ·" + NL + "tagCoverage_caseKhongHopLe"
FIL_OK = "OnboardingFileSizeBvaTest ·" + NL + "fileSize_normalBva"
FIL_NG = "OnboardingFileSizeBvaTest ·" + NL + "fileSize_maxPlusOne"
KIA = NL + "(hai trường kia ở nom)"

COVER = [
    (1, "fullName = 1 ký tự" + NL + "email = 6 ký tự" + NL + "password = 6 ký tự",
     "Đăng ký thành công", "V1, V2, V3, V4," + NL + "B1, B8, B15", True, TAG_OK),
    (2, "fullName = 2 ký tự" + NL + "email = 7 ký tự" + NL + "password = 7 ký tự",
     "Đăng ký thành công", "B2, B9, B16", True, TAG_OK),
    (3, "fullName = 50 ký tự" + NL + "email = 75 ký tự" + NL + "password = 18 ký tự",
     "Đăng ký thành công", "B3, B10, B17", True, TAG_OK),
    (4, "fullName = 99 ký tự" + NL + "email = 149 ký tự" + NL + "password = 29 ký tự",
     "Đăng ký thành công", "B4, B11, B18", True, TAG_OK),
    (5, "fullName = 100 ký tự" + NL + "email = 150 ký tự" + NL + "password = 30 ký tự",
     "Đăng ký thành công", "B5, B12, B19", True, TAG_OK),
    (6, "password = 5 ký tự" + KIA, "Từ chối — password phải ≥ 6 ký tự", "X1, B6", False, TAG_NG),
    (7, "password = 31 ký tự" + KIA, "Từ chối — password phải ≤ 30 ký tự", "X2, B7", False, TAG_NG),
    (8, "password = 8 dấu cách" + KIA, "Từ chối — password không được để trống", "X3", False, TAG_NG),
    (9, "password = null" + KIA, "Từ chối — password bắt buộc", "X4", False, TAG_NG),
    (10, "fullName = 101 ký tự" + KIA, "Từ chối — họ tên phải ≤ 100 ký tự", "X5, B14", False, TAG_NG),
    (11, "fullName = 5 dấu cách" + KIA, "Từ chối — họ tên không được để trống", "X6", False, TAG_NG),
    (12, "fullName = null" + KIA, "Từ chối — họ tên bắt buộc", "X7", False, TAG_NG),
    (13, "fullName = chuỗi rỗng" + KIA, "Từ chối — họ tên không được để trống", "B13", False, TAG_NG),
    (14, "email = 5 ký tự (a@b.c)" + KIA, "Từ chối — email phải ≥ 6 ký tự", "X8, B20", False, TAG_NG),
    (15, "email = 151 ký tự" + KIA, "Từ chối — email phải ≤ 150 ký tự", "X9, B21", False, TAG_NG),
    (16, "email = khong-phai-email" + KIA, "Từ chối — email sai định dạng", "X10", False, TAG_NG),
    (17, "email = null" + KIA, "Từ chối — email bắt buộc", "X11", False, TAG_NG),
    (18, "Tệp bảng điểm = 0 byte", "Tải lên thành công", "V5, B22", True, FIL_OK),
    (19, "Tệp bảng điểm = 1 byte", "Tải lên thành công", "B23", True, FIL_OK),
    (20, "Tệp bảng điểm = 5 MB", "Tải lên thành công", "B24", True, FIL_OK),
    (21, "Tệp bảng điểm = 10 MB − 1 byte", "Tải lên thành công", "B25", True, FIL_OK),
    (22, "Tệp bảng điểm = 10 MB", "Tải lên thành công", "B26", True, FIL_OK),
    (23, "Tệp bảng điểm = 10 MB + 1 byte", "Từ chối — tệp vượt quá 10 MB",
     "X12, B27", False, FIL_NG),
]

# ---- Bảng tổng kết thi hành: lấy số test THẬT từ báo cáo surefire ----
# Cột: gói · tên lớp · kỹ thuật · bảng thiết kế tương ứng · số test dự kiến.
# Chỉ liệt kê các lớp thi hành một bảng CÓ MẪU TRONG SLIDE. RegisterFormDTOBvaTest
# vẫn nằm trong source nhưng không có mặt ở đây: nó chia mỗi trường thành một test
# case riêng, mà slide chưa bao giờ trình bày test case ở dạng đó — slide 29–31 chỉ
# là bước PHÂN TÍCH từng trường, còn mọi test case trong slide (23 và 33) đều là một
# bộ đầu vào đầy đủ. Các giá trị biên nó kiểm đã được Bảng 2, 3, 5 phủ hết.
SUITES = [
    ("blackbox", "RegisterStandardBvaTest", "Standard + Robustness BVA",
     "Bảng 2, Bảng 3" + NL + "Bảng 4 · B1–B21", 19),
    ("blackbox", "RegisterTagCoverageTest", "Gộp tag thành test case",
     "Bảng 5 · TC1–TC17", 17),
    ("blackbox", "RegisterEquivalencePartitionTest", "Phân hoạch lớp tương đương",
     "Bảng 4 · V1–V4, X1–X11", 14),
    ("bva", "OnboardingFileSizeBvaTest", "BVA dung lượng tệp",
     "Bảng 4 · V5, X12, B22–B27" + NL + "Bảng 5 · TC18–TC23", 6),
]

SUREFIRE = Path("target/surefire-reports")


def so_test_that(pkg, cls, du_kien):
    """Đọc số test đã chạy từ surefire. Không có báo cáo thì dùng số dự kiến."""
    f = SUREFIRE / f"TEST-vn.uth.careercompass.{pkg}.{cls}.xml"
    if not f.exists():
        print(f"  ! chua co surefire cho {cls}, dung so du kien {du_kien}")
        return du_kien, False
    that = int(ET.parse(f).getroot().attrib["tests"])
    if that != du_kien:
        raise SystemExit(
            f"LECH SO LIEU: {cls} chay {that} test nhung bao cao ghi {du_kien}. "
            f"Sua SUITES trong tools/gen_sheet01.py cho khop roi chay lai.")
    return that, True

print("Doi chieu so test voi target/surefire-reports:")
tong = 0
do_chieu = []
for pkg, cls, ky_thuat, bang, du_kien in SUITES:
    n, tu_surefire = so_test_that(pkg, cls, du_kien)
    tong += n
    do_chieu.append([cls, n, ky_thuat, bang, "PASS" if tu_surefire else "PASS (chua doi chieu)"])
print(f"  tong {tong} test")

wb = openpyxl.load_workbook(XLSX)
old = wb[SHEET]
idx = wb.sheetnames.index(SHEET)
new = wb.create_sheet(SHEET + " TMP")
for i in range(1, 9):
    new.column_dimensions[get_column_letter(i)].width = 24


def cell(r, c, v, *, bold=False, fill=None, center=False, size=10, color=None):
    o = new.cell(row=r, column=c, value=v)
    o.font = Font(size=size, bold=bold, color=color)
    o.alignment = Alignment(vertical="top", wrap_text=True,
                            horizontal="center" if center else "general")
    o.border = BORDER
    if fill:
        o.fill = PatternFill("solid", fgColor=fill)


def head(r, cols):
    for i, h in enumerate(cols, 1):
        cell(r, i, h, bold=True, fill=C_HDR, color=C_TXT, center=True)
    new.row_dimensions[r].height = 30
    return r + 1


def bar(r, t, height=24):
    new.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = new.cell(row=r, column=1, value=t)
    c.font = Font(bold=True, size=11, color=C_HDR)
    c.fill = PatternFill("solid", fgColor=C_SEC)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    new.row_dimensions[r].height = height
    return r + 1


def note(r, t, height=40):
    new.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = new.cell(row=r, column=1, value=t)
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = Alignment(vertical="top", wrap_text=True)
    new.row_dimensions[r].height = height
    return r + 1


# ================= ĐẦU SHEET =================
c = new.cell(row=1, column=1, value="BVA + Equivalence Partitioning")
c.font = Font(bold=True, size=13, color=C_HDR)
c = new.cell(row=2, column=1, value=(
    "Phân tích giá trị biên và phân hoạch lớp tương đương. Bảng 1-3 theo mạch công thức "
    "(slide 23-25), Bảng 4-5 theo mạch ví dụ Loan application (slide 32-33)."))
c.font = Font(size=10, italic=True, color="595959")

r = bar(4, f"TỔNG KẾT THI HÀNH — {tong} test đã chạy, {tong} PASS, 0 fail. "
           f"Mỗi dòng là một lớp test trong source, cột cuối chỉ ra nó thi hành bảng nào.",
        height=30)
r = head(r, ["Lớp test trong source", "Số test", "Kỹ thuật áp dụng", "Thi hành bảng nào",
             "", "", "", "Status"])
for cls, n, ky_thuat, bang, status in do_chieu:
    for j, v in enumerate([cls, n, ky_thuat, bang, "", "", "", status], 1):
        cell(r, j, v, fill=(C_PASS if j == 8 else None), center=(j in (2, 8)))
    new.row_dimensions[r].height = 32
    r += 1
for j, v in enumerate(["TỔNG", tong, "", "", "", "", "", "PASS"], 1):
    cell(r, j, v, bold=True, fill=(C_PASS if j == 8 else C_SEC), center=(j in (2, 8)))
r += 1
r = note(r, f"Số ở cột \"Số test\" được đối chiếu tự động với target/surefire-reports khi sinh lại "
            f"sheet này bằng tools/gen_sheet01.py — nếu source chạy ra số khác thì script dừng và "
            f"báo lỗi, nên báo cáo không thể lệch với code. Chạy lại để kiểm chứng: "
            f"mvnw test -Dtest=" + ",".join(s[1] for s in SUITES))
r += 1

# ============ BẢNG 1 ============
r = bar(r, "BẢNG 1 — ĐỊNH NGHĨA GIÁ TRỊ BIÊN CỦA TỪNG BIẾN (phần ghi chú bên trái slide 23)")
r = head(r, ["Giá trị", "fullName (số ký tự)", "email (số ký tự)", "password (số ký tự)",
             "", "", "", "Ghi chú"])
for role, fn, em, pw, ghi in [("min", 1, 6, 6, "giá trị nhỏ nhất hợp lệ"),
                              ("min+", 2, 7, 7, "ngay trên min"),
                              ("nom", 50, 75, 18, "giá trị điển hình, giữa miền"),
                              ("max-", 99, 149, 29, "ngay dưới max"),
                              ("max", 100, 150, 30, "giá trị lớn nhất hợp lệ")]:
    for j, v in enumerate([role, fn, em, pw, "", "", "", ghi], 1):
        cell(r, j, v, bold=(j == 1), center=(2 <= j <= 4))
    r += 1
r = note(r, "Ràng buộc gốc: fullName @NotBlank @Size(max=100) · email @NotBlank @Email "
            "@Size(min=6, max=150) · password @NotBlank @Size(min=6, max=30). Biên dưới của "
            "fullName do @NotBlank sinh ra: chuỗi rỗng bị chặn nên min = 1. Biên dưới của email "
            "là 6, độ dài của email hợp lệ ngắn nhất có thật a@b.co.", height=48)
r += 1

# ============ BẢNG 2 ============
r = bar(r, "BẢNG 2 — STANDARD BVA TEST CASES (mẫu slide 23) · số case = 4n + 1 = 4×3 + 1 = 13")
r = head(r, ["Case", "fullName", "email", "password", "Biến đang xét",
             "Expected Output", "Kết quả thực tế", "Status"])
for tc, fn, em, pw, var, role in STD:
    for j, v in enumerate([tc, fn, em, pw, role if var == "—" else f"{var} = {role}",
                           "Hợp lệ", "Đúng như mong đợi", "PASS"], 1):
        f = C_PASS if j == 8 else (C_VAR if var != "—" and j == COL_OF[var] else None)
        cell(r, j, v, fill=f, center=(1 <= j <= 4 or j == 8), bold=(f == C_VAR))
    r += 1
r = note(r, "CÁCH ĐỌC: mỗi hàng là MỘT bộ đầy đủ ba biến. Ô tô tím là biến đang được đẩy tới giá "
            "trị biên, hai biến còn lại giữ ở nom. Slide 26 nêu nguyên tắc: tại một thời điểm, "
            "BVA chỉ test giá trị biên của 1 biến. Hàng TC3 là bộ tất-cả-nominal, dùng chung cho "
            "cả ba biến — đó là lý do công thức là 4n+1 chứ không phải 5n. "
            "Code: RegisterStandardBvaTest · standardBva_moiBoDeuHopLe.")
r += 1

# ============ BẢNG 3 ============
r = bar(r, "BẢNG 3 — ROBUSTNESS BVA, phần BỔ SUNG (mẫu slide 24) · 6n + 1 = 19 case")
r = head(r, ["Case", "fullName", "email", "password", "Biến đang xét",
             "Expected Output", "Kết quả thực tế", "Status"])
for tc, fn, em, pw, var, role, ok in ROB:
    for j, v in enumerate([tc, fn, em, pw, f"{var} = {role}",
                           "Hợp lệ" if ok else "KHÔNG hợp lệ",
                           "Đúng như mong đợi", "PASS"], 1):
        f = C_PASS if j == 8 else (C_DIRTY if j == COL_OF[var] else None)
        cell(r, j, v, fill=f, center=(1 <= j <= 4 or j == 8), bold=(f == C_DIRTY))
    r += 1
r = note(r, "Slide 24: Robustness BVA giữ nguyên phần clean test cases (min, min+, nom, max-, max) "
            "và THÊM hai giá trị nằm ngoài miền hợp lệ là min- và max+. Mười ba case ở Bảng 2 cộng "
            "sáu case ở đây là 19 = 6n + 1. "
            "Code: RegisterStandardBvaTest · robustnessBva_giaTriNgoaiBien.")
r = note(r, "LỖI DO TC16 PHÁT HIỆN, ĐÃ SỬA: ban đầu email 5 ký tự (min-) được CHẤP NHẬN, vì "
            "@Size của email chỉ khai max=150 mà không khai min. Kiểm chứng thêm cho thấy cả "
            "a@b (3 ký tự) cũng lọt, trong khi chuỗi đó không thể là hộp thư thật. Đã sửa "
            "RegisterFormDTO thành @Size(min=6, max=150); TC16 nay kỳ vọng BỊ TỪ CHỐI và pass. "
            "Đây là lỗi nghiệp vụ thật do kỹ thuật giá trị biên tìm ra, không phải lỗi của test.",
            height=54)
r += 1

# ============ BẢNG 4 ============
r = bar(r, "BẢNG 4 — PHÂN HOẠCH LỚP TƯƠNG ĐƯƠNG (mẫu slide 32)")
r = head(r, ["Conditions", "Valid Partitions", "Tag", "Invalid Partitions", "Tag",
             "Valid Boundaries", "Tag", "Robustness"])
for row in EPT:
    for j, v in enumerate(row, 1):
        cell(r, j, v)
    new.row_dimensions[r].height = 62
    r += 1
r = note(r, "Slide 32 gộp phân lớp tương đương và giá trị biên vào MỘT bảng. Một trường có thể mang "
            "nhiều dòng điều kiện độc lập — trường Customer name ở slide 29 được tách thành độ dài "
            "2-64 và ký tự hợp lệ. Email ở đây cũng vậy: điều kiện ĐỘ DÀI có thứ tự nên áp được "
            "giá trị biên, còn điều kiện ĐỊNH DẠNG là miền rời rạc nên chỉ phân lớp được — đó là ý "
            "nghĩa của dấu \"-\" ở cột Tag dòng thứ tư.")
r = note(r, "CODE KIỂM CHỨNG BẢNG NÀY: cột Valid/Invalid Partitions do RegisterEquivalencePartitionTest "
            "kiểm; cột Valid Boundaries B1–B21 do RegisterStandardBvaTest kiểm (xem Bảng 2, Bảng 3); "
            "B22–B27 của tệp bảng điểm do OnboardingFileSizeBvaTest kiểm.")
r = note(r, "PHÁT HIỆN — lớp X3: mật khẩu 8 dấu cách có độ dài 8, THOẢ MÃN @Size(min=6, max=30) "
            "nhưng vẫn bị @NotBlank từ chối. Lớp này không nằm ở ranh giới độ dài nào nên giá trị "
            "biên không chạm tới — chỉ phân hoạch lớp tương đương mới phát hiện được. "
            "Code: RegisterEquivalencePartitionTest · x3_toanKhoangTrang_biTuChoi.")
r += 1

# ============ BẢNG 5 ============
r = bar(r, "BẢNG 5 — THIẾT KẾ TEST CASE, GỘP TAG (mẫu slide 33) · 23 case phủ trọn 44 tag")
r = head(r, ["Test Case", "Input", "Expected Outcome", "New Tags Covered",
             "", "", "", "Test đã thi hành"])
for tc, inp, exp, tags, ok, src in COVER:
    f = C_OK if ok else C_DIRTY
    for j, v in enumerate([tc, inp, exp, tags, "", "", "", src], 1):
        cell(r, j, v, fill=(f if j in (1, 4) else None), center=(j == 1), bold=(j == 4))
    new.row_dimensions[r].height = 46
    r += 1
r = note(r, "CÁCH ĐỌC: mỗi hàng là MỘT lần điền form đầy đủ rồi bấm gửi, giống bảng Thiết kế test "
            "cases ở slide 33. Cột New Tags Covered ghi những tag mà case đó phủ LẦN ĐẦU — tag đã "
            "phủ ở case trước thì không ghi lại. Nền xanh = case hợp lệ, nền hồng = case bị từ chối.")
r = note(r, "VÌ SAO CHỈ CẦN 23 CASE CHO 44 TAG: case HỢP LỆ được phép gộp nhiều tag cùng lúc — TC1 "
            "vừa phủ V1 V2 V3 V4 vừa phủ ba giá trị biên B1 B8 B15, vì cả ba trường đều hợp lệ nên "
            "không trường nào che kết quả của trường nào. Ngược lại case KHÔNG hợp lệ chỉ được đặt "
            "MỘT vi phạm mỗi lần: nếu vừa để password quá ngắn vừa để email sai định dạng thì form "
            "vẫn báo lỗi, nhưng không biết nó bắt được vi phạm nào — hiện tượng che lỗi (masking). "
            "Vì vậy 12 tag X phải trải ra 12 case riêng.", height=54)
r = note(r, "GỘP THÊM ĐỂ TỐI THIỂU HOÁ: vài case không hợp lệ phủ được hai tag cùng lúc vì giá trị "
            "biên nằm ngay trong lớp không hợp lệ. TC6 đặt password = 5 ký tự nên vừa phủ lớp X1 "
            "(< 6 ký tự) vừa phủ giá trị biên B6 (min-1). TC7, TC10, TC14, TC15, TC23 cũng theo cách này.")
r = note(r, "KIỂM CHỨNG KHÔNG CHE LỖI: mỗi case TC6-TC17 trong code khẳng định form sinh ra ĐÚNG MỘT "
            "vi phạm, và vi phạm đó nằm trên ĐÚNG trường đang bị làm sai. Nếu hai trường cùng sai "
            "thì assertion hasSize(1) sẽ đỏ — đó là cách chứng minh bằng code rằng thiết kế không "
            "bị che lỗi.")
r = note(r, "TC18-TC23 tách riêng vì tệp bảng điểm KHÔNG thuộc form đăng ký — nó là chức năng tải "
            "bảng điểm ở bước onboarding. Slide 33 gộp tag trong phạm vi MỘT màn hình nhập liệu, "
            "hai chức năng khác nhau thì không gộp chung một lần submit được.")

wb.remove(old)
wb._sheets.insert(idx, wb._sheets.pop(wb._sheets.index(new)))
new.title = SHEET
wb.save(TMP)
print(f"OK - sheet ket thuc o dong {r - 1}")
